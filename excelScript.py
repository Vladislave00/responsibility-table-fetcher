import openpyxl
import openpyxl.styles
import dbalch


# Метод для форматирования границ таблицы
def set_border(ws, cell_range, need_to_thick, need_to_thick_up, need_to_thick_down):
    thin = openpyxl.styles.Side(border_style="thin", color="000000")
    thick = openpyxl.styles.Side(border_style="thick", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            if cell == row[len(row) - 1]:
                cell.border = openpyxl.styles.Border(top=thin, left=thin, right=thick, bottom=thin)
                cell.fill = openpyxl.styles.PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            else:
                cell.border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
        if len(need_to_thick) != 0:
            if row == ws[need_to_thick[0]]:
                for cell in row:
                    cell.border = openpyxl.styles.Border(top=thick, left=thick, right=thick, bottom=thick)
                need_to_thick.pop(0)
        if len(need_to_thick_up) != 0:
            if row == ws[need_to_thick_up[0]]:
                for cell in row:
                    if cell == row[len(row) - 1]:
                        cell.border = openpyxl.styles.Border(top=thick, left=thin, right=thick, bottom=thin)
                    else:
                        cell.border = openpyxl.styles.Border(top=thick, left=thin, right=thin, bottom=thin)
                need_to_thick_up.pop(0)
        if len(need_to_thick_down) != 0:
            if row == ws[need_to_thick_down[0]]:
                for cell in row:
                    if cell == row[len(row) - 1]:
                        cell.border = openpyxl.styles.Border(top=thin, left=thin, right=thick, bottom=thick)
                    else:
                        cell.border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thick)

                need_to_thick_down.pop(0)


# Метод для формирования Excel таблицы
def make_excel():
    # Получаем метрики из базы данных
    data = dbalch.Database.get_data()
    # Создаем книгу и заполняем ее
    wb = openpyxl.Workbook()

    # добавляем новый лист
    wb.create_sheet(title='Зоны ответственности', index=0)

    # получаем лист, с которым будем работать
    sheet = wb['Зоны ответственности']
    ws = wb.active

    bold_italic_font = openpyxl.styles.Font(bold=True, italic=True)
    italic_font = openpyxl.styles.Font(italic=True)

    # формирование шапки таблицы
    header_data = ["№", "№", "Показатели", "Имя сотрудника", "Должность"]
    ws.append(["Зона ответственности экспертов по вводу сведений в системе оценки деятельности заведующих кафедрами Московского политехнического университета"])
    cell = ws.cell(row=ws.max_row, column=1)
    cell.font = bold_italic_font
    ws.merge_cells(start_column=1, start_row=1, end_row=1, end_column=5)
    ws.append(header_data)

    current_row = 2
    current_category = ""
    current_subname = ""

    # Заполнение строк данными из базы данных
    for row in data:
        if current_category != row.activity_name:
                current_category = row.activity_name
                ws.append([row.activity_name])
                cell = ws.cell(row=ws.max_row, column=1)
                cell.font = bold_italic_font
                current_row+=1
                ws.merge_cells(start_column=1, start_row=current_row, end_row=current_row, end_column=5)

        if row.subname != None:
            if row.subname != "" and current_subname != row.subname:
                current_subname = row.subname
                ws.append([row.subname])
                cell = ws.cell(row=ws.max_row, column=1)
                cell.font = italic_font
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                current_row+=1
                ws.merge_cells(start_column=1, start_row=current_row, end_row=current_row, end_column=5)

        ws.append([row.number, row.letter, row.indicator, row.employee_name, row.employee_post])
        current_row+=1
    
    for row in ws.iter_rows(min_row=1, max_row=current_row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='center', vertical='center')

    # Выставление ширины колонок
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20

    thin = openpyxl.styles.Side(border_style="thin", color="000000")
    thick = openpyxl.styles.Side(border_style="medium", color="000000")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)

    # Сохранение файла
    wb.save('Зоны ответственности.xlsx')
    print("Файл успешно сохранен")


make_excel()